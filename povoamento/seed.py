import sqlite3
import re
from openpyxl import load_workbook

'''
    Funções úteis -------------------------------------------------------
'''
def sanitize_input(text):
    """
    Sanitiza uma entrada de texto.
    - Remove espaços extras no início e no fim.
    - Substitui múltiplos espaços consecutivos por um único espaço.

    Args:
        text (str): Entrada de texto a ser sanitizada.

    Returns:
        str | None: Texto sanitizado ou None se a entrada for vazia.
    """
    if not text and text != 0:  # Verifica se o texto é vazio ou None e diferente de 0
        return None

    if isinstance(text, str):  # Confirma que a entrada é uma string
        sanitized_text = text.strip()  # Remove espaços extras nas extremidades
        sanitized_text = ' '.join(sanitized_text.split())  # Substitui múltiplos espaços por um único
        return sanitized_text

    return text  # Retorna a entrada original caso não seja string


def sanitize_values(values):
    """
    Sanitiza os valores para uma tabela específica.

    Args:
        table_name (str): Nome da tabela.
        values (list): Lista contendo os valores a serem sanitizados.

    Returns:
        list: Lista de valores sanitizados.
    """
    sanitized_values = []
    for v in values:
        sanitized_values.append(sanitize_input(v))
        
    return sanitized_values


def validate_identifiers(table_name, *columns):
    """
    Valida se os nomes da tabela e das colunas são identificadores válidos.

    Args:
        table_name (str): Nome da tabela.
        *columns (str): Nomes das colunas a serem validadas.

    Raises:
        ValueError: Se algum nome for inválido.
    """
    if not table_name.isidentifier() or not all(col.isidentifier() for col in columns):
        raise ValueError("Nomes de tabela ou coluna inválidos")


def record_exists(cursor, table_name, column_name, value):
    """
    Verifica se um registro existe em uma tabela para um determinado valor em uma coluna.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar consultas no banco de dados.
        table_name (str): Nome da tabela onde a busca será realizada.
        column_name (str): Nome da coluna para filtrar os registros.
        value (str): Valor a ser verificado.

    Returns:
        tuple | None: O registro encontrado ou None se não houver correspondência.
    """
    value = sanitize_values([value])[0]
    query = f"SELECT * FROM {table_name} WHERE {column_name} = ?"
    cursor.execute(query, (value,))
    return cursor.fetchone()


def record_exists_with_two_values(cursor, table_name, column1, column2, values):
    """
    Verifica se um registro existe em uma tabela com base em dois valores.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar consultas no banco de dados.
        table_name (str): Nome da tabela onde a busca será realizada.
        column1 (str): Nome da primeira coluna para filtrar os registros.
        column2 (str): Nome da segunda coluna para filtrar os registros.
        value1 (str | None): Valor da primeira coluna (pode ser None).
        value2 (str): Valor da segunda coluna.

    Returns:
        tuple | None: O registro encontrado ou None se não houver correspondência.
    """
    # Sanitiza os valores de entrada
    [value1_sanitized, value2_sanitized] = sanitize_values(values)

    # Define a consulta SQL com base nas condições dos valores
    if table_name in ["Distritos", "Municipios"] and not value1_sanitized:
        # Se a primeira coluna deve ser NULL
        query = f"SELECT * FROM {table_name} WHERE {column1} IS NULL AND {column2} = ?"
        parameters = (value2_sanitized,)
    else:
        # Busca normal com ambos os valores
        query = f"SELECT * FROM {table_name} WHERE {column1} = ? AND {column2} = ?"
        parameters = (value1_sanitized, value2_sanitized)

    # Executa a consulta e retorna o resultado
    cursor.execute(query, parameters)
    return cursor.fetchone()

def handle_entity_values(value_list, index):
    """
    Manipula entradas específicas para o caso de 'Entidades'.

    Args:
        value_list (list): Lista de valores separados por '|'.
        index (int): Índice atual da iteração.

    Returns:
        tuple: Uma tupla contendo:
            - Lista de valores divididos por " - ".
            - Boolean indicando se o próximo valor deve ser ignorado.
    """
    # Verifica se o próximo valor deve ser combinado com o atual
    if index != len(value_list) - 1 and len(value_list[index + 1].split(" - ", 1)) == 1:
        combined_value = value_list[index] + "|" + value_list[index + 1]
        return combined_value.split(" - ", 1), True
    return value_list[index].split(" - ", 1), False


'''
    Inserção generica de valores na tabela ------------------------------
'''
def insert_one_value(cursor, table_name, column_name, value):
    """
    Insere um único valor em uma tabela e retorna o ID da nova linha.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        table_name (str): Nome da tabela.
        column_name (str): Nome da coluna onde o valor será inserido.
        value (str): Valor a ser inserido.

    Returns:
        int: ID da linha inserida.
    """
    query = f"""
        INSERT INTO {table_name} ({column_name}) 
        VALUES (?);
    """
    cursor.execute(query, (value,))
    return cursor.lastrowid


def get_or_insert_single_value(cursor, table_name, column_name, value):
    """
    Verifica se um valor existe em uma tabela e retorna seu ID.
    Caso o valor não exista, insere-o na tabela e retorna o ID da nova linha.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        table_name (str): Nome da tabela.
        column_name (str): Nome da coluna a ser verificada/inserida.
        value (str): Valor a ser verificado ou inserido.

    Returns:
        int: ID do registro existente ou recém-inserido.
    """
    # Verifica se os nomes da tabela e da coluna são válidos
    validate_identifiers(table_name, column_name)

    # Sanitiza o valor de entrada
    value = sanitize_input(value)

    # Verifica a existência do valor na tabela
    found_row = record_exists(cursor, table_name, column_name, value)

    if found_row: # se o valor já foi adicionado retorna o ID do registro encontrado
        return found_row[0]  
    elif value: # se ele não foi adicionado e é não nulo, nos o adicionamos na bd
        return insert_one_value(cursor, table_name, column_name, value)


def insert_two_values(cursor, table_name, column1, column2, values):
    """
    Insere dois valores em uma tabela e retorna o ID da nova linha.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        table_name (str): Nome da tabela.
        column1 (str): Nome da primeira coluna.
        column2 (str): Nome da segunda coluna.
        value1 (str | None): Valor da primeira coluna (pode ser None).
        value2 (str): Valor da segunda coluna.

    Returns:
        int: ID da linha inserida.
    """
    query = f"""
        INSERT INTO {table_name} ({column1}, {column2}) 
        VALUES (?, ?);
    """
    cursor.execute(query, (values[0], values[1]))
    return cursor.lastrowid


def get_or_insert_double_value(cursor, table_name, column1_name, column2_name, values, check_function=record_exists):
    """
    Verifica se dois valores existem em uma tabela e retorna seu ID.
    Caso os valores não existam, insere-os na tabela e retorna o ID da nova linha.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        table_name (str): Nome da tabela.
        column1_name (str): Nome da primeira coluna.
        column2_name (str): Nome da segunda coluna.
        values (list): Lista contendo dois valores a serem verificados/inseridos.
        check_function (function): Função para verificar a existência dos valores, por padrão, verifica apenas um valor e não a tupla.

    Returns:
        int: ID do registro existente ou recém-inserido.
    """
    # Verifica se os nomes da tabela e das colunas são válidos
    validate_identifiers(table_name, column1_name, column2_name)

    # Sanitiza os valores de entrada
    values = sanitize_values(values)

    # Verifica a existência dos valores na tabela
    if check_function == record_exists:
        found_row = check_function(cursor, table_name, column1_name, values[0])
    else:
        found_row = check_function(cursor, table_name, column1_name, column2_name, values)

    if found_row:
        return found_row[0]  # Retorna o ID do registro encontrado
    elif (values[0] and values[1]) or (table_name in ["Distritos", "Municipios"] and values[1]):
        return insert_two_values(cursor, table_name, column1_name, column2_name, values)

'''
    Inserção especifica de valores na tabela ----------------------------
'''
def insert_or_get_contract_id(cursor, contract_data, related_ids):
    """
    Insere ou retorna o ID de um contrato na tabela "Contratos".

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        contract_data (dict): Dicionário contendo os dados do contrato, com as chaves:
                              - idcontrato, objectoContrato, dataPublicacao,
                                dataCelebracaoContrato, precoContratual,
                                ProcedimentoCentralizado, prazoExecucao.
        related_ids (dict): Dicionário contendo IDs relacionados ao contrato, com as chaves:
                            - idProcedimento, idAdjudicante.

    Returns:
        int: ID do contrato existente ou recém-inserido.
    """
    # Sanitiza todos os valores do dicionário de dados do contrato
    sanitized_data = {key: sanitize_input(value) for key, value in contract_data.items()}

    # Verifica se o contrato já existe
    existing_row = record_exists(cursor, "Contratos", "idContrato", sanitized_data["idcontrato"])
    if existing_row:
        return existing_row[0]

    # Prepara o comando SQL para inserir um novo contrato
    insert_query = """
        INSERT INTO Contratos (
            idContrato, objetoContrato, dataPublicacao, dataCelebracaoContrato,
            precoContratual, procedimentoCentralizado, prazoExecucao,
            idProcedimento, idAdjudicante
        ) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
    """

    # Executa a inserção do novo contrato
    cursor.execute(
        insert_query,
        (
            sanitized_data["idcontrato"],
            sanitized_data["objectoContrato"],
            sanitized_data["dataPublicacao"],
            sanitized_data["dataCelebracaoContrato"],
            sanitized_data["precoContratual"],
            sanitized_data["ProcedimentoCentralizado"],
            sanitized_data["prazoExecucao"],
            related_ids["idProcedimento"],
            related_ids["idAdjudicante"][0],
        ),
    )

    # Retorna o ID do novo contrato inserido
    return cursor.lastrowid


def get_ids_from_multiple_values(cursor, data, operation_type, table_name, column1, column2=None, existence_check=record_exists):
    """
    Obtém ou insere IDs com base em múltiplos valores separados por '|' e retorna uma lista de IDs.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        data (str): String contendo múltiplos valores separados por '|'.
        operation_type (function): Função de operação a ser usada (single ou double value).
        table_name (str): Nome da tabela.
        column1 (str): Nome da primeira coluna.
        column2 (str, optional): Nome da segunda coluna. Necessário para operações de double value.
        existence_check (function): Função para verificar a existência de valores.

    Returns:
        list: Lista de IDs obtidos ou recém-inseridos.
    """
    value_list = data.split("|")
    ids = []
    skip_next = False

    for i in range(len(value_list)):
        if skip_next:
            skip_next = False
            continue

        if operation_type == get_or_insert_single_value:
            # Operação para uma única coluna
            entity_id = get_or_insert_single_value(cursor, table_name, column1, value_list[i])

        elif operation_type == get_or_insert_double_value:
            # Operação para duas colunas
            if table_name == "CPVs":
                value_pair = value_list[i].split(" - ", 1)
            elif table_name == "Entidades":
                value_pair, skip_next = handle_entity_values(value_list, i)
            else:
                raise ValueError("Operação não suportada para esta tabela")

            if len(value_pair) != 2: # Valor inválido ou incompleto
                continue

            entity_id = get_or_insert_double_value(cursor, table_name, column1, column2, value_pair, existence_check)

        ids.append(entity_id)
    return ids


def associate_contract_with_values(cursor, table_name, column1, column2, contract_id, value_ids):
    """
    Associa múltiplos valores a um contrato na tabela especificada.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        table_name (str): Nome da tabela.
        column1 (str): Nome da primeira coluna (ID do contrato).
        column2 (str): Nome da segunda coluna (ID relacionado).
        contract_id (int): ID do contrato.
        value_ids (list): Lista de IDs a serem associados ao contrato.
    """
    for value_id in value_ids:
        get_or_insert_double_value(
            cursor,
            table_name,
            column1,
            column2,
            [contract_id, value_id],
            record_exists_with_two_values
        )


def check_fundamentation_existence(cursor, fundamentation_data):
    """
    Verifica a existência de uma fundamentação na tabela 'fundamentacoes'.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        fundamentation_data (dict): Dicionário com as chaves:
                                    - artigo, n, alinea, subalinea, referenciaLegislativa.

    Returns:
        tuple: Linha encontrada na tabela, ou None se não existir.
    """
    query = """
        SELECT * FROM fundamentacoes 
        WHERE artigo = ? AND numero = ? AND alinea = ? AND subalinea = ? AND referenciaLegislativa = ?;
    """
    parameters = (
        fundamentation_data["artigo"],
        fundamentation_data["numero"],
        fundamentation_data["alinea"],
        fundamentation_data["subalinea"],
        fundamentation_data["referenciaLegislativa"],
    )
    cursor.execute(query, parameters)
    return cursor.fetchone()


def insert_fundamentation(cursor, fundamentation_data):
    """
    Insere ou retorna o ID de uma fundamentação com base nos dados fornecidos.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        fundamentation_data (dict): Dicionário contendo os campos:
            - artigo, numero, alinea, subalinea, referenciaLegislativa.

    Returns:
        int: ID da fundamentação existente ou recém-inserida.
    """
    # Verifica se já existe uma fundamentação com os dados fornecidos
    found_row = check_fundamentation_existence(cursor, fundamentation_data)
    if found_row:
        return found_row[0]

    # Insere uma nova fundamentação
    query = """
        INSERT INTO fundamentacoes (artigo, numero, alinea, subalinea, referenciaLegislativa)
        VALUES (?, ?, ?, ?, ?);
    """
    parameters = (
        fundamentation_data["artigo"],
        fundamentation_data["numero"],
        fundamentation_data["alinea"],
        fundamentation_data["subalinea"],
        fundamentation_data["referenciaLegislativa"],
    )
    cursor.execute(query, parameters)
    return cursor.lastrowid


def process_fundamentations(cursor, raw_data):
    """
    Processa e insere múltiplas fundamentações a partir de uma string de entrada.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        raw_data (str): String contendo fundamentações separadas por " e ".

    Returns:
        list: Lista de IDs das fundamentações processadas.
    """
    if not raw_data:
        return []

    fundamentations = raw_data.split(" e ")
    ids = []

    for f in fundamentations:
        fundamentation_data = {
            "artigo": sanitize_input(extract_pattern(f, r"artigo\s(\d+)[.º]")),
            "numero": sanitize_input(extract_pattern(f, r"n.º\s(\d),")),
            "alinea": sanitize_input(extract_pattern(f, r"alínea\s([a-zA-Z]+)[)]")),
            "subalinea": sanitize_input(extract_pattern(f, r"subalínea\s([a-zA-Z]+)[)]")),
            "referenciaLegislativa": sanitize_input(
                extract_pattern(f, r"do\s(.+)") or
                extract_pattern(f, r"da\s(.+)") or
                extract_pattern(f, r"dos\s(.+)")
            ),
        }

        if any(fundamentation_data.values()):  # Verifica se há ao menos um valor válido
            ids.append(insert_fundamentation(cursor, fundamentation_data))

    return ids

def extract_pattern(text, pattern):
    """
    Extrai o primeiro grupo correspondente ao padrão fornecido em uma string.

    Args:
        text (str): Texto no qual buscar o padrão.
        pattern (str): Expressão regular para busca.

    Returns:
        str: Valor encontrado ou None.
    """
    match = re.search(pattern, text, re.IGNORECASE)
    return match.group(1) if match else None


'''
    processamento de dados do contrato ----------------------------
'''
def process_location(cursor, location_string):
    """
    Processa uma string de localização no formato "país,distrito,município" (separados por "|")
    e retorna uma lista de IDs dos municípios correspondentes.

    Args:
        cursor (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        location_string (str): String contendo localizações separadas por "|", 
                               onde cada localização segue o formato "país,distrito,município".

    Returns:
        list: Lista de IDs dos municípios processados.
    """
    # Divide a string em diferentes localizações
    locations = location_string.split("|")
    # Divide cada localização nos componentes: país, distrito, município
    parsed_locations = [loc.split(",") for loc in locations]

    municipality_ids = []

    for location in parsed_locations:
        # Obtém o país (obrigatório)
        country = location[0]

        # Obtém distrito e município, se existirem
        district = location[1] if len(location) > 1 else None
        municipality = location[2] if len(location) > 2 else None

        # Insere/obtém o ID do país
        country_id = get_or_insert_single_value(cursor, "Paises", "pais", country)

        # Insere/obtém o ID do distrito (relacionado ao país)
        district_id = get_or_insert_double_value( cursor, "Distritos", "distrito", "idPais", [district, country_id], record_exists_with_two_values)

        # Insere/obtém o ID do município (relacionado ao distrito)
        municipality_id = get_or_insert_double_value( cursor, "Municipios", "municipio", "idDistrito", [municipality, district_id], record_exists_with_two_values)

        # Adiciona o ID do município à lista
        municipality_ids.append(municipality_id)

    return municipality_ids


def process_cpvs(cur, data):
    """
    Processa os códigos CPV (Classificação de Produtos e Serviços) a partir dos dados fornecidos
    e retorna uma lista de IDs dos CPVs correspondentes, inserindo-os se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        data (dict): Dicionário contendo os dados de CPVs, com a chave "cpv" contendo a lista de códigos.

    Returns:
        list: Lista de IDs dos CPVs processados e inseridos no banco de dados.
    """
    return get_ids_from_multiple_values(cur, data, get_or_insert_double_value, "CPVs", "codigoCPV", "descricaoCPV")
    

def process_procedure(cur, data):
    """
    Processa o tipo de procedimento do contrato e retorna o ID do tipo de procedimento correspondente,
    inserindo-o no banco de dados se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        data (dict): Dicionário contendo o tipo de procedimento do contrato.

    Returns:
        int: ID do tipo de procedimento processado e inserido no banco de dados.
    """
    #TODO talvez normalizar
    return get_or_insert_single_value(cur, "TiposProcedimentos", "procedimento", data)


def process_contract_types(cur, data):
    """
    Processa os tipos de contrato fornecidos e retorna uma lista de IDs dos tipos de contrato correspondentes,
    inserindo-os no banco de dados se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        data (list): Lista de tipos de contrato.

    Returns:
        list: Lista de IDs dos tipos de contrato processados e inseridos no banco de dados.
    """
    return get_ids_from_multiple_values(cur, data, get_or_insert_single_value, "TiposContratos", "tipo")


def process_entities(cur, data):
    """
    Processa as entidades fornecidas (identificadas pelo NIF) e retorna uma lista de IDs das entidades correspondentes,
    inserindo-as no banco de dados se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        data (list): Lista de NIFs das entidades.

    Returns:
        list: Lista de IDs das entidades processadas e inseridas no banco de dados.
    """
    return get_ids_from_multiple_values(cur, data, get_or_insert_double_value, "Entidades", "nif", "entidade", record_exists_with_two_values)


def process_agreement(cur, data):
    """
    Processa os acordos quadro (identificados pelo NIF) e retorna uma lista de IDs dos acordos correspondentes,
    inserindo-os no banco de dados se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        data (list): Lista de NIFs dos acordos.

    Returns:
        list: Lista de IDs dos acordos processados e inseridos no banco de dados.
    """
    return get_ids_from_multiple_values(cur, data, get_or_insert_single_value, "DescrAcordoQuadro", "descricao")

def create_contract(cur, data, related_ids):
    """
    Cria um novo contrato no banco de dados e retorna o ID do contrato criado.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        data (dict): Dicionário contendo os dados do contrato.
        related_ids (dict): Dicionário contendo os IDs relacionados ao contrato.

    Returns:
        int: ID do contrato criado.
    """
    #TODO - procedimento centralizado ta indo pra bd como 1 ou 2 , não como 0 ou 1, ver o motivo
    values_contrato = {"idcontrato": data["idcontrato"], "objectoContrato": data["objectoContrato"], "dataPublicacao": data["dataPublicacao"], "dataCelebracaoContrato": data["dataCelebracaoContrato"], "precoContratual": data["precoContratual"], "ProcedimentoCentralizado": (1 if data["ProcedimentoCentralizado"] == "Sim" else 0 ), "prazoExecucao": data["prazoExecucao"]}
    return insert_or_get_contract_id(cur, values_contrato, related_ids)


def associate_agreement(cur, idContrato, data):
    """
    Associa um contrato a um acordo quadro específico, inserindo o acordo quadro se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        idContrato (int): ID do contrato a ser associado ao acordo quadro.
        data (list): Lista de dados dos acordos quadro a serem associados.
    """
    idAcordoQuadro = process_agreement(cur, data)
    associate_contract_with_values(cur, "AcordoQuadroContratos", "idContrato", "idAcordoQuadro", idContrato, idAcordoQuadro)


def associate_adjudicators(cur, idContrato, data):
    """
    Associa um contrato a um ou mais adjudicatários, inserindo as entidades correspondentes se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        idContrato (int): ID do contrato a ser associado aos adjudicatários.
        data (list): Lista de entidades (adjudicatários) a serem associadas, identificadas pelo NIF.
    """
    idAdjudicatarios = process_entities(cur, data)
    associate_contract_with_values(cur, "Adjudicatarios", "idContrato", "idEntidade", idContrato, idAdjudicatarios)


def associate_cpvs(cur, idContrato, data):
    """
    Associa um contrato a códigos CPV específicos, inserindo os códigos correspondentes se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        idContrato (int): ID do contrato a ser associado aos CPVs.
        data (list): Lista de códigos CPV a serem associados.
    """
    codigoCPV = process_cpvs(cur, data)
    associate_contract_with_values(cur, "CPVContratos", "idContrato", "codigoCPV", idContrato, codigoCPV)


def associate_contract_types(cur, idContrato, data):
    """
    Associa um contrato a tipos de contrato específicos, inserindo os tipos correspondentes se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        idContrato (int): ID do contrato a ser associado aos tipos de contrato.
        data (list): Lista de tipos de contrato a serem associados.
    """
    idTipoContrato = process_contract_types(cur, data)
    associate_contract_with_values(cur, "ClassificacaoContratos", "idContrato", "idTipoContrato", idContrato, idTipoContrato)


def associate_locations(cur, idContrato, data):
    """
    Associa um contrato a locais de execução, inserindo os municípios correspondentes se necessário.

    Args:
        cur (sqlite3.Cursor): Cursor para executar comandos no banco de dados.
        idContrato (int): ID do contrato a ser associado aos locais de execução.
        data (str): String contendo os municípios de execução, no formato "país,distrito,município".
    """
    idMunicipios = process_location(cur, data)
    associate_contract_with_values(cur, "LocaisDeExecucao", "idContrato", "idMunicipio", idContrato, idMunicipios)


def associate_contract_fundamentations(cur, idContrato, data):
    idFundamentacao = process_fundamentations(cur, data)
    associate_contract_with_values(cur, "FundamentacaoContratos", "idContrato", "idFundamentacao", idContrato, idFundamentacao)


def process_contract_data(cur, row_data):
    """
    Processa e insere um contrato e suas associações no banco de dados.
    """
    related_ids = {}

    # Obtenção de IDs que são associados ao contrato
    related_ids["idProcedimento"] = process_procedure(cur, row_data["tipoprocedimento"])
    related_ids["idAdjudicante"] = process_entities(cur, row_data["adjudicante"])

    # Criação do contrato
    idContrato = create_contract(cur, row_data, related_ids)

    # Associações específicas
    if row_data["DescrAcordoQuadro"] != "NULL":
        associate_agreement(cur, idContrato, row_data["DescrAcordoQuadro"])
    if row_data.get("adjudicatarios"):
        associate_adjudicators(cur, idContrato, row_data["adjudicatarios"])
    associate_locations(cur, idContrato, row_data["localExecucao"])
    associate_cpvs(cur, idContrato, row_data["cpv"])
    associate_contract_types(cur, idContrato, row_data["tipoContrato"])
    #associate_contract_fundamentations(cur, idContrato, row_data["fundamentacao"])
    '''
        TODO - ajeitar fundamentação os dados de referencia Legislativa estao aparecendo como null então não consegue adicionar na bd
    '''


def add_dataset(sheet): 
    """
    Esta função recebe uma planilha do Excel, extrai os dados e os adiciona em um banco de dados SQLite.

    A função assume que a planilha está estruturada com os dados começando na segunda linha, sendo que
    a primeira linha contém os cabeçalhos (nomes das colunas). Para cada linha de dados, ela cria um dicionário
    associando cada coluna ao seu respectivo valor, e em seguida, adiciona os dados no banco de dados.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): A planilha do Excel a ser processada.
    """
    # Conectando ao banco de dados SQLite
    con = sqlite3.connect('../contratos_publicos.db')
    cur = con.cursor()

    # Obtendo os cabeçalhos da primeira linha
    headers = [cell.value for cell in sheet[1]] 

    # Iterando sobre as linhas de dados (começando da segunda linha)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Criando um dicionário associando cada cabeçalho ao valor da respectiva célula na linha
        row_data = dict(zip(headers, row))     
        
        # TODO: Melhorar a checagem de existência do contrato fora desta função, 
        # garantindo que contratos duplicados não sejam adicionados.
        process_contract_data(cur, row_data)

    # Commit das alterações no banco de dados
    con.commit()
    print("Adicionado na BD")
    
    # Fechando a conexão com o banco de dados
    con.close()

# Carregando o arquivo Excel
workbook = load_workbook(filename='../dataset/ContratosPublicos2024.xlsx')

# Selecionando a planilha ativa do arquivo
sheet = workbook.active

# Adicionando os dados da planilha ao banco de dados
add_dataset(sheet)
